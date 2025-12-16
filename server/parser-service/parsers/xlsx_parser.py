import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from typing import Dict, Any, List
from .base_parser import BaseParser

class XLSXParser(BaseParser):
    def parse(self, file_path: str) -> Dict[str, Any]:
        result = self.create_result_structure()
        
        try:
            # Загрузка с data_only=True для значений, False для формул
            wb_data = load_workbook(file_path, data_only=True)
            wb_formulas = load_workbook(file_path, data_only=False)
            
            result['metadata'] = {
                'type': 'xlsx',
                'sheets': wb_data.sheetnames,
                'sheet_count': len(wb_data.sheetnames),
                'title': wb_data.properties.title or '',
                'author': wb_data.properties.creator or '',
                'created': str(wb_data.properties.created) if wb_data.properties.created else '',
                'modified': str(wb_data.properties.modified) if wb_data.properties.modified else '',
                'last_modified_by': wb_data.properties.lastModifiedBy or '',
                'subject': wb_data.properties.subject or '',
                'keywords': wb_data.properties.keywords or '',
            }
            
            all_text = []
            structure = []
            tables = []
            charts = []
            
            for sheet_name in wb_data.sheetnames:
                sheet_data = wb_data[sheet_name]
                sheet_formulas = wb_formulas[sheet_name]
                
                # Базовая информация о листе
                sheet_info = {
                    'sheet': sheet_name,
                    'max_row': sheet_data.max_row,
                    'max_column': sheet_data.max_column,
                    'rows': []
                }
                
                # Извлечение данных построчно
                for row_idx, row in enumerate(sheet_data.iter_rows(values_only=False), start=1):
                    row_data = []
                    row_text = []
                    
                    for cell_idx, cell in enumerate(row, start=1):
                        if cell.value is not None:
                            # Получение формулы если есть
                            formula_cell = sheet_formulas.cell(row_idx, cell_idx)
                            formula = formula_cell.value if isinstance(formula_cell.value, str) and formula_cell.value.startswith('=') else None
                            
                            cell_info = {
                                'value': str(cell.value),
                                'formula': formula,
                                'data_type': cell.data_type,
                                'number_format': cell.number_format,
                                'coordinate': cell.coordinate,
                            }
                            
                            # Информация о стиле
                            if cell.font:
                                cell_info['style'] = {
                                    'bold': cell.font.bold,
                                    'italic': cell.font.italic,
                                    'underline': cell.font.underline,
                                    'color': str(cell.font.color.rgb) if cell.font.color and hasattr(cell.font.color, 'rgb') else None,
                                    'size': cell.font.size
                                }
                            
                            if cell.fill:
                                cell_info['fill'] = {
                                    'pattern': cell.fill.patternType,
                                    'fgColor': str(cell.fill.fgColor.rgb) if cell.fill.fgColor and hasattr(cell.fill.fgColor, 'rgb') else None
                                }
                            
                            row_data.append(cell_info)
                            row_text.append(str(cell.value))
                    
                    if row_data:
                        sheet_info['rows'].append(row_data)
                        all_text.append(' | '.join(row_text))
                
                structure.append(sheet_info)
                
                # Извлечение таблицы с использованием pandas
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                    
                    table_data = {
                        'sheet': sheet_name,
                        'headers': df.columns.tolist(),
                        'rows': [[str(cell) if pd.notna(cell) else '' for cell in row] for row in df.values.tolist()],
                        'row_count': len(df),
                        'col_count': len(df.columns),
                        'has_na': df.isna().any().any(),
                        'dtypes': {col: str(dtype) for col, dtype in df.dtypes.items()}
                    }
                    tables.append(table_data)
                except Exception as e:
                    pass
                
                # Извлечение графиков
                if hasattr(sheet_data, '_charts'):
                    for chart in sheet_data._charts:
                        charts.append({
                            'sheet': sheet_name,
                            'type': chart.__class__.__name__,
                            'title': chart.title if hasattr(chart, 'title') else ''
                        })
            
            result['content']['text'] = '\n'.join(all_text)
            result['content']['structure'] = structure
            result['content']['tables'] = tables
            result['content']['charts'] = charts
            
            # Статистика
            result['metadata']['table_count'] = len(tables)
            result['metadata']['chart_count'] = len(charts)
            result['metadata']['cell_count'] = sum(sheet['max_row'] * sheet['max_column'] for sheet in structure)
            result['metadata']['word_count'] = len(' '.join(all_text).split())
            
            wb_data.close()
            wb_formulas.close()
            
        except Exception as e:
            result['metadata']['error'] = str(e)
        
        return result
